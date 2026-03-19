import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from extract import extract_all
from transform import transform

def load_to_db(df):
    conn = sqlite3.connect("ecommerce.db")
    df.to_sql("orders", conn, if_exists="replace", index=False)
    print(f"✅ Saved {len(df)} rows to ecommerce.db")
    conn.close()


def export_excel(df):
    # --- Calculate KPIs ---
    total_revenue = df['price'].sum()
    top_categories = df.groupby('product_category_name_english')['price'].sum().nlargest(5)
    monthly_orders = df.groupby(df['order_purchase_timestamp'].dt.to_period('M')).size()
    avg_delivery = (df['order_delivered_customer_date'] - df['order_purchase_timestamp']).dt.days.mean()

    # --- Build Excel ---
    wb = openpyxl.Workbook()

    # Sheet 1 — Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1['A1'] = "Olist KPI Report"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A2'] = "Metric"
    ws1['B2'] = "Value"
    for cell in ['A2', 'B2']:
        ws1[cell].font = Font(bold=True, color="FFFFFF")
        ws1[cell].fill = PatternFill("solid", fgColor="2E86AB")

    ws1['A3'] = "Total Revenue"
    ws1['B3'] = f"${total_revenue:,.2f}"
    ws1['A4'] = "Avg Delivery Days"
    ws1['B4'] = f"{avg_delivery:.1f} days"
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 20

    # Sheet 2 — Top Categories
    ws2 = wb.create_sheet("Top Categories")
    ws2['A1'] = "Category"
    ws2['B1'] = "Revenue"
    for cell in ['A1', 'B1']:
        ws2[cell].font = Font(bold=True, color="FFFFFF")
        ws2[cell].fill = PatternFill("solid", fgColor="2E86AB")
    for i, (cat, val) in enumerate(top_categories.items(), start=2):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2, value=round(val, 2))
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15

    # Sheet 3 — Monthly Orders
    ws3 = wb.create_sheet("Monthly Orders")
    ws3['A1'] = "Month"
    ws3['B1'] = "Orders"
    for cell in ['A1', 'B1']:
        ws3[cell].font = Font(bold=True, color="FFFFFF")
        ws3[cell].fill = PatternFill("solid", fgColor="2E86AB")
    for i, (month, count) in enumerate(monthly_orders.items(), start=2):
        ws3.cell(row=i, column=1, value=str(month))
        ws3.cell(row=i, column=2, value=count)
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 15

    wb.save("olist_report.xlsx")
    print("✅ Excel report saved → olist_report.xlsx")

if __name__ == "__main__":
    data = extract_all("archive")
    master_df = transform(data)
    load_to_db(master_df)
    export_excel(master_df)
