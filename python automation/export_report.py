import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def clean_data(df):
    df = df.copy()
    df['date'] = pd.to_datetime(df['date'])
    df['returned'] = df['returned'].astype(int)
    df = df[df['units_sold'] >= 1]
    df['revenue'] = df['units_sold'] * df['unit_price']
    print(f"✅ Cleaning done — {len(df)} rows remaining")
    return df

def analyze(df):
    results = {}
    results['total_revenue'] = df['revenue'].sum()
    results['best_product'] = df.groupby('product')['revenue'].sum().idxmax()
    results['top_reps'] = df.groupby('sales_rep')['revenue'].sum().nlargest(3)
    results['monthly_trend'] = df.groupby(df['date'].dt.to_period('M'))['revenue'].sum()
    results['return_rate'] = df['returned'].mean() * 100
    return results

def export_report(df):
    kpis = analyze(df)
    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws1 = wb.active
    ws1.title = "Summary"

    # Title
    ws1['A1'] = "Sales KPI Report"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1.merge_cells('A1:B1')

    # Headers
    ws1['A2'] = "Metric"
    ws1['B2'] = "Value"
    for cell in ['A2', 'B2']:
        ws1[cell].font = Font(bold=True, color="FFFFFF")
        ws1[cell].fill = PatternFill("solid", fgColor="2E86AB")
        ws1[cell].alignment = Alignment(horizontal="center")

    # KPI rows
    rows = [
        ("Total Revenue", f"₹{kpis['total_revenue']:,.0f}"),
        ("Best Product", kpis['best_product']),
        ("Return Rate", f"{kpis['return_rate']:.1f}%"),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws1[f'A{i}'] = label
        ws1[f'B{i}'] = value
        if i % 2 == 0:
            for col in ['A', 'B']:
                ws1[f'{col}{i}'].fill = PatternFill("solid", fgColor="D9EAF7")

    # Column widths
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 20

    # --- Sheet 2: Monthly Trend ---
    ws2 = wb.create_sheet("Monthly Trend")
    ws2['A1'] = "Month"
    ws2['B1'] = "Revenue"
    for cell in ['A1', 'B1']:
        ws2[cell].font = Font(bold=True, color="FFFFFF")
        ws2[cell].fill = PatternFill("solid", fgColor="2E86AB")
        ws2[cell].alignment = Alignment(horizontal="center")

    for i, (month, value) in enumerate(kpis['monthly_trend'].items()):
        ws2.cell(row=i+2, column=1, value=str(month))
        ws2.cell(row=i+2, column=2, value=value)

    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 15

    # --- Sheet 3: Top Reps ---
    ws3 = wb.create_sheet("Top Reps")
    ws3['A1'] = "Sales Rep"
    ws3['B1'] = "Revenue"
    for cell in ['A1', 'B1']:
        ws3[cell].font = Font(bold=True, color="FFFFFF")
        ws3[cell].fill = PatternFill("solid", fgColor="2E86AB")

    for i, (rep, value) in enumerate(kpis['top_reps'].items()):
        ws3.cell(row=i+2, column=1, value=rep)
        ws3.cell(row=i+2, column=2, value=value)

    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 15

    wb.save("sales_report.xlsx")
    print("✅ Report saved → sales_report.xlsx")

# --- Run everything ---
df_raw = pd.read_csv("sales_raw.csv")
df_clean = clean_data(df_raw)
export_report(df_clean)